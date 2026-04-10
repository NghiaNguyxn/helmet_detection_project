import io
from datetime import datetime, date
from motor.motor_asyncio import AsyncIOMotorCollection
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

async def export_violations_to_excel(
    collection: AsyncIOMotorCollection,
    start_date: date,
    end_date: date
) -> io.BytesIO:
    """Export errors to an Excel file, return BytesIO."""
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    cursor = collection.find({"timestamp": {"$gte": start_dt, "$lte": end_dt}}).sort("timestamp", -1)
    violations = await cursor.to_list(length=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations Report"

    # Header
    headers = ["Timestamp", "Image URL", "Total Violations", "Detections (class, confidence)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data rows
    for row_idx, v in enumerate(violations, 2):
        ts = v["timestamp"].strftime("%Y-%m-%d %H:%M:%S") if isinstance(v["timestamp"], datetime) else str(v["timestamp"])
        image_url = v.get("image_url", "")
        total_violations = v.get("total_violations", 0)
        detections = v.get("detections", [])
        det_text = "; ".join([f"{d.get('class_name')} ({d.get('confidence',0):.2f})" for d in detections])

        ws.cell(row=row_idx, column=1, value=ts)
        ws.cell(row=row_idx, column=2, value=image_url)
        ws.cell(row=row_idx, column=3, value=total_violations)
        ws.cell(row=row_idx, column=4, value=det_text)

    # Auto-adjust columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_len + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output