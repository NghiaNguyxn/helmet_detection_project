from datetime import datetime
import asyncio
import csv
import io
import json
import zipfile
from bson import ObjectId
from bson.errors import InvalidId
from motor.motor_asyncio import AsyncIOMotorCollection
from pymongo import ReturnDocument
import numpy as np
import openpyxl
import requests
from openpyxl.styles import Font, Alignment, PatternFill
import logging

from SourceCode.BE.app.services.upload_service import upload_image_to_cloudinary
from SourceCode.BE.app.schemas.helmet_schema import Detection, ViolationHistoryResponse, ViolationRecord
from SourceCode.BE.app.core.websocket_manager import manager
from SourceCode.BE.app.enums.violation_status import ViolationStatus
from SourceCode.BE.app.enums.rejection_reason import RejectionReason
from SourceCode.BE.app.models.user import UserDB
from SourceCode.BE.app.schemas.user_schema import ReviewUser

logger = logging.getLogger(__name__)

def _build_violation_query(
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False,
    status: ViolationStatus | str | None = None
) -> dict:
    """Helper to build MongoDB query for violations"""
    
    query = {}
    
    if start_date or end_date:
        query["timestamp"] = {}
        if start_date:
            query["timestamp"]["$gte"] = start_date
        if end_date:
            query["timestamp"]["$lte"] = end_date
            
    if min_violations is not None:
        query["total_violations"] = {"$gte": min_violations}
        
    if only_violations:
        current_min = query.get("total_violations", {}).get("$gte", 0)
        query["total_violations"] = {"$gte": max(1, current_min)}

    if status and status != "all":
        status_value = status.value if isinstance(status, ViolationStatus) else status
        query["status"] = status_value
        
    return query

async def get_violation_history(
    db_collection: AsyncIOMotorCollection, 
    page: int, 
    limit: int,
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False,
    status: ViolationStatus | str | None = None,
    sort_by: str = "timestamp",
    order: str = "desc"
) -> ViolationHistoryResponse:
    """
    Get paginated and filtered violation history from MongoDB Atlas.
    """

    # 1. Xây dựng bộ lọc (Query Filter)
    query = _build_violation_query(start_date, end_date, min_violations, only_violations, status)

    # 2. Tính toán phân trang
    skip = (page - 1) * limit

    # 3. Lấy tổng số bản ghi khớp với bộ lọc
    total_count = await db_collection.count_documents(query)

    # 4. Truy vấn dữ liệu với Sorting
    sort_direction = -1 if order == "desc" else 1
    cursor = db_collection.find(query).sort(sort_by, sort_direction).skip(skip).limit(limit)

    violations: list[ViolationRecord] = []
    async for doc in cursor:
        violations.append(_serialize_violation_doc(doc))

    # 5. Tạo response
    return ViolationHistoryResponse(
        total=total_count,
        page=page,
        limit=limit,
        data=violations,
    )

async def delete_violation(db_collection: AsyncIOMotorCollection, violation_id: str) -> bool:
    """Delete a violation record by ID"""

    object_id = _to_object_id(violation_id)
    if object_id is None:
        return False

    result = await db_collection.delete_one({"_id": object_id})
    if result.deleted_count > 0:
        # Broadcast để các tab khác xóa cache/UI
        await manager.broadcast({
            "event": "delete_violation",
            "data": {"id": violation_id}
        })

    return result.deleted_count > 0

async def save_violation_backtask(
        annotated_frame: np.ndarray, 
        violation_count: int, 
        all_detections: list[Detection], 
        db_collection: AsyncIOMotorCollection,
        camera_context: dict | None = None,
    ):
    """Background task for saving violation record"""

    try:
        cloud_url = await upload_image_to_cloudinary(annotated_frame)
        logger.info(f"Image uploaded to Cloudinary: {cloud_url}")

        timestamp_obj = datetime.now()

        violation_doc = ViolationRecord(
            timestamp=timestamp_obj,
            image_url=cloud_url,
            total_violations=violation_count,
            detections=[d for d in all_detections if d.class_id == 1],
            camera_id=camera_context.get("camera_id") if camera_context else None,
            camera_code=camera_context.get("camera_code") if camera_context else None,
            camera_name=camera_context.get("camera_name") if camera_context else None,
            camera_location=camera_context.get("camera_location") if camera_context else None,
            camera_source_type=camera_context.get("camera_source_type") if camera_context else None,
            is_demo=bool(camera_context.get("is_demo")) if camera_context else False,
            status=ViolationStatus.PENDING,
            reviewed_by=None,
            reviewed_at=None,
            review_note=None,
            rejection_reason=None,
        )
        # Chuyển đổi sang dict nhưng loại bỏ trường 'id' vì MongoDB sẽ tự tạo '_id'
        inserted_doc = violation_doc.model_dump(exclude={'id'})
        await db_collection.insert_one(inserted_doc)
        
        # Gán ID vừa tạo vào doc trước khi gửi qua WebSocket
        if "_id" in inserted_doc:
            violation_doc.id = str(inserted_doc["_id"])
            
        logger.info(f"[Background] Saved violation to Cloud with ID: {violation_doc.id}")

        # Gửi dữ liệu qua WebSocket tới các client
        payload = {
            "event": "new_violation",
            "data": violation_doc.model_dump(mode='json')
        }
        await manager.broadcast(payload)
        logger.info(f"[Background] Broadcasted new violation to WebSockets")
    except Exception as e:
        logger.error(f"[Background] Failed to save history: {e}")

async def confirm_violation(
        db_collection: AsyncIOMotorCollection, 
        violation_id: str, 
        review_note: str | None, 
        reviewer: UserDB
) -> ViolationRecord | None:
    """
    Confirm a violation with optional review note, then broadcast the update via WebSocket.
    """

    object_id = _to_object_id(violation_id)
    if object_id is None:
        return None
    
    reviewed_at = datetime.now()
    update_doc = await db_collection.find_one_and_update(
        {"_id": object_id},
        {
            "$set": {
                "status": ViolationStatus.CONFIRMED.value,
                "reviewed_by": _reviewer_snapshot(reviewer),
                "reviewed_at": reviewed_at,
                "review_note": review_note,
                "rejection_reason": None
            }
        },
        return_document = ReturnDocument.AFTER
    )
    if not update_doc:
        return None
    
    violation_record = _serialize_violation_doc(update_doc)
    await manager.broadcast({
        "event": "review_violation",
        "data": _review_event_payload(violation_record)
    })
    logger.info(f"Violation {violation_id} confirmed by {reviewer.username}. Broadcasted update to WebSockets.")

    return violation_record

async def reject_violation(
        db_collection: AsyncIOMotorCollection, 
        violation_id: str, 
        rejection_reason: RejectionReason, 
        review_note: str | None, 
        reviewer: UserDB
) -> ViolationRecord | None:
    """
     Reject a violation with reason and optional note, then broadcast the update via WebSocket.
    """

    object_id = _to_object_id(violation_id)
    if object_id is None:
        return None
    
    reviewed_at = datetime.now()
    update_doc = await db_collection.find_one_and_update(
        {"_id": object_id},
        {
            "$set": {
                "status": ViolationStatus.REJECTED.value,
                "reviewed_by": _reviewer_snapshot(reviewer),
                "reviewed_at": reviewed_at,
                "review_note": review_note,
                "rejection_reason": rejection_reason.value
            }
        },
        return_document = ReturnDocument.AFTER
    )
    if not update_doc:
        return None

    violation_record = _serialize_violation_doc(update_doc)
    await manager.broadcast({
        "event": "review_violation",
        "data": _review_event_payload(violation_record)
    })
    logger.info(f"Violation {violation_id} rejected by {reviewer.username}. Broadcasted update to WebSockets.")

    return violation_record

async def export_violations_to_excel(
    db_collection: AsyncIOMotorCollection,
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False,
    status: ViolationStatus | str | None = None
) -> io.BytesIO:
    """Export violations to Excel with professional styling using openpyxl"""
    
    query = _build_violation_query(start_date, end_date, min_violations, only_violations, status)
    cursor = db_collection.find(query).sort("timestamp", -1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lịch sử vi phạm"
    
    # 1. Định dạng tiêu đề
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    
    headers = [
        "ID", 
        "Timestamp", 
        "Violation Count", 
        "Average Confidence", 
        "Image URL",
        "Violation Status",
        "Reviewed By",
        "Reviewed At",
        "Review Note",
        "Rejection Reason"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 2. Thêm dữ liệu vào bảng
    row_idx = 2
    async for doc in cursor:
        detections = doc.get("detections", [])
        avg_conf = sum(d.get("confidence", 0) for d in detections) / len(detections) if detections else 0
        
        ws.cell(row=row_idx, column=1, value=str(doc["_id"]))
        ws.cell(row=row_idx, column=2, value=doc.get("timestamp").isoformat() if doc.get("timestamp") else "")
        ws.cell(row=row_idx, column=3, value=doc.get("total_violations", 0))
        ws.cell(row=row_idx, column=4, value=round(avg_conf, 2))
        ws.cell(row=row_idx, column=5, value=doc.get("image_url", ""))
        ws.cell(row=row_idx, column=6, value=doc.get("status", ""))
        ws.cell(row=row_idx, column=7, value=doc.get("reviewed_by", {}).get("username", "") if doc.get("reviewed_by") else "")
        ws.cell(row=row_idx, column=8, value=doc.get("reviewed_at").isoformat() if doc.get("reviewed_at") else "")
        ws.cell(row=row_idx, column=9, value=doc.get("review_note", ""))
        ws.cell(row=row_idx, column=10, value=doc.get("rejection_reason", ""))
        row_idx += 1

    # 3. Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

async def export_feedback_dataset(
    db_collection: AsyncIOMotorCollection,
    start_date: datetime = None,
    end_date: datetime = None,
    status: ViolationStatus | str | None = "all",
    rejection_reason: RejectionReason | str | None = None,
    include_images: bool = True,
    limit: int = 500,
) -> io.BytesIO:
    """Export reviewed violations as an AI feedback dataset ZIP."""

    query = _build_violation_query(
        start_date=start_date,
        end_date=end_date,
    )

    status_value = status.value if isinstance(status, ViolationStatus) else status
    if status_value == "all":
        query["status"] = {
            "$in": [
                ViolationStatus.CONFIRMED.value,
                ViolationStatus.REJECTED.value,
            ]
        }
    elif status_value:
        query["status"] = status_value

    if rejection_reason and rejection_reason != "all":
        reason_value = rejection_reason.value if isinstance(rejection_reason, RejectionReason) else rejection_reason
        query["status"] = ViolationStatus.REJECTED.value
        query["rejection_reason"] = reason_value

    cursor = db_collection.find(query).sort("timestamp", -1).limit(limit)
    output = io.BytesIO()

    manifest_rows = []
    image_jobs = []
    detection_files = []

    async for doc in cursor:
        record_id = str(doc["_id"])
        status_value = doc.get("status", "")
        reason_value = doc.get("rejection_reason", "")
        detections = doc.get("detections", [])
        avg_conf = sum(d.get("confidence", 0) for d in detections) / len(detections) if detections else 0
        image_file = f"images/{status_value or 'unknown'}/{record_id}.jpg"
        detections_file = f"detections/{record_id}.json"

        manifest_rows.append({
            "id": record_id,
            "timestamp": doc.get("timestamp").isoformat() if doc.get("timestamp") else "",
            "status": status_value,
            "feedback_label": _feedback_label(status_value, reason_value),
            "rejection_reason": reason_value,
            "reviewed_by": doc.get("reviewed_by", {}).get("username", "") if doc.get("reviewed_by") else "",
            "reviewed_at": doc.get("reviewed_at").isoformat() if doc.get("reviewed_at") else "",
            "review_note": doc.get("review_note", ""),
            "total_violations": doc.get("total_violations", 0),
            "detection_count": len(detections),
            "average_confidence": round(avg_conf, 4),
            "image_url": doc.get("image_url", ""),
            "image_file": image_file if include_images else "",
            "detections_file": detections_file,
        })
        detection_files.append((detections_file, _feedback_detection_payload(doc, record_id)))

        if include_images and doc.get("image_url"):
            image_jobs.append((doc.get("image_url"), image_file))

    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        manifest_buffer = io.StringIO()
        fieldnames = [
            "id",
            "timestamp",
            "status",
            "feedback_label",
            "rejection_reason",
            "reviewed_by",
            "reviewed_at",
            "review_note",
            "total_violations",
            "detection_count",
            "average_confidence",
            "image_url",
            "image_file",
            "detections_file",
        ]
        writer = csv.DictWriter(manifest_buffer, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(manifest_rows)
        archive.writestr("manifest.csv", manifest_buffer.getvalue())

        readme = (
            "AI Feedback Dataset\n"
            "\n"
            "This export is built from reviewed violation records.\n"
            f"Maximum exported records for this request: {limit}.\n"
            "- confirmed records are positive violation examples.\n"
            "- rejected records are feedback examples grouped by rejection reason.\n"
            "- image_file points to the downloaded evidence image when include_images=true.\n"
            "- detections_file points to per-record bounding boxes, classes, and confidence values.\n"
            "\n"
            "Note: current evidence images may be annotated images, not raw camera frames.\n"
        )
        archive.writestr("README.txt", readme)

        for detections_file, detection_payload in detection_files:
            archive.writestr(
                detections_file,
                json.dumps(detection_payload, ensure_ascii=False, indent=2, default=str)
            )

        if include_images:
            for image_url, image_file in image_jobs:
                image_bytes = await asyncio.to_thread(_download_image_bytes, image_url)
                if image_bytes:
                    archive.writestr(image_file, image_bytes)

    output.seek(0)
    return output

def _to_object_id(violation_id: str) -> ObjectId | None:
    try:
        return ObjectId(violation_id)
    except (InvalidId, TypeError):
        return None
    
def _serialize_violation_doc(doc: dict) -> ViolationRecord:
    record = {**doc, "id": str(doc["_id"])}
    return ViolationRecord(**record)

def _reviewer_snapshot(reviewer: UserDB) -> dict:
    return ReviewUser(
        id=reviewer.id,
        username=reviewer.username,
        role=reviewer.role
    ).model_dump(mode="json")

def _review_event_payload(violation: ViolationRecord) -> dict:
    return {
        "id": violation.id,
        "status": violation.status.value,
        "reviewed_by": violation.reviewed_by.model_dump() if violation.reviewed_by else None,
        "reviewed_at": violation.reviewed_at.isoformat() if violation.reviewed_at else None,
        "review_note": violation.review_note,
        "rejection_reason": violation.rejection_reason.value if violation.rejection_reason else None
    }

def _feedback_label(status: str, rejection_reason: str | None) -> str:
    if status == ViolationStatus.CONFIRMED.value:
        return "confirmed_violation"
    if status == ViolationStatus.REJECTED.value:
        return rejection_reason or "rejected"
    return status or "unknown"

def _feedback_detection_payload(doc: dict, record_id: str) -> dict:
    return {
        "id": record_id,
        "timestamp": doc.get("timestamp").isoformat() if doc.get("timestamp") else "",
        "status": doc.get("status", ""),
        "feedback_label": _feedback_label(doc.get("status", ""), doc.get("rejection_reason")),
        "rejection_reason": doc.get("rejection_reason"),
        "image_url": doc.get("image_url", ""),
        "detections": doc.get("detections", []),
    }

def _download_image_bytes(image_url: str) -> bytes | None:
    try:
        response = requests.get(image_url, timeout=15)
        response.raise_for_status()
        return response.content
    except requests.RequestException as exc:
        logger.warning("Failed to download feedback image %s: %s", image_url, exc)
        return None
