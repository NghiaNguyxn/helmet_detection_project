from datetime import datetime
import io
from bson import ObjectId
from motor.motor_asyncio import AsyncIOMotorCollection
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from SourceCode.BE.app.services.upload_service import upload_image_to_cloudinary
from SourceCode.BE.app.schemas.helmet_schema import Detection, ViolationHistoryResponse, ViolationRecord

def _build_violation_query(
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False
) -> dict:
    """Helper to build MongoDB query for violations"""
    query = {}
    
    if start_date or end_date:
        query["timestamp"] = {}
        if start_date:
            query["timestamp"]["$gte"] = start_date
        if end_date:
            query["timestamp"]["$lte"] = end_date
            
    if min_violations is not None:
        query["total_violations"] = {"$gte": min_violations}
        
    if only_violations:
        current_min = query.get("total_violations", {}).get("$gte", 0)
        query["total_violations"] = {"$gte": max(1, current_min)}
        
    return query

async def get_violation_history(
    db_collection: AsyncIOMotorCollection, 
    page: int, 
    limit: int,
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False,
    sort_by: str = "timestamp",
    order: str = "desc"
) -> ViolationHistoryResponse:
    """
    Get paginated and filtered violation history from MongoDB Atlas.
    """

    # 1. Xây dựng bộ lọc (Query Filter)
    query = _build_violation_query(start_date, end_date, min_violations, only_violations)

    # 2. Tính toán phân trang
    skip = (page - 1) * limit

    # 3. Lấy tổng số bản ghi khớp với bộ lọc
    total_count = await db_collection.count_documents(query)

    # 4. Truy vấn dữ liệu với Sorting
    sort_direction = -1 if order == "desc" else 1
    cursor = db_collection.find(query).sort(sort_by, sort_direction).skip(skip).limit(limit)

    violations: list[ViolationRecord] = []
    async for doc in cursor:
        doc["id"] = str(doc["_id"])
        violations.append(ViolationRecord(**doc))

    # 5. Tạo response
    return ViolationHistoryResponse(
        total=total_count,
        page=page,
        limit=limit,
        data=violations,
    )

async def delete_violation(db_collection: AsyncIOMotorCollection, violation_id: str) -> bool:
    """Delete a violation record by ID"""
    
    result = await db_collection.delete_one({"_id": ObjectId(violation_id)})
    return result.deleted_count > 0

async def export_violations_to_excel(
    db_collection: AsyncIOMotorCollection,
    start_date: datetime = None,
    end_date: datetime = None,
    min_violations: int = None,
    only_violations: bool = False
) -> io.BytesIO:
    """Export violations to Excel with professional styling using openpyxl"""
    
    query = _build_violation_query(start_date, end_date, min_violations, only_violations)
    cursor = db_collection.find(query).sort("timestamp", -1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lịch sử vi phạm"
    
    # 1. Header Styling
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    
    headers = ["ID", "Thời gian", "Số người vi phạm", "Confidence TB", "URL Ảnh"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 2. Add Data
    row_idx = 2
    async for doc in cursor:
        detections = doc.get("detections", [])
        avg_conf = sum(d.get("confidence", 0) for d in detections) / len(detections) if detections else 0
        
        ws.cell(row=row_idx, column=1, value=str(doc["_id"]))
        ws.cell(row=row_idx, column=2, value=doc.get("timestamp").strftime("%Y-%m-%d %H:%M:%S") if doc.get("timestamp") else "N/A")
        ws.cell(row=row_idx, column=3, value=doc.get("total_violations", 0))
        ws.cell(row=row_idx, column=4, value=round(avg_conf, 2))
        ws.cell(row=row_idx, column=5, value=doc.get("image_url", ""))
        row_idx += 1

    # 3. Auto-adjust columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

async def save_violation_backtask(
        annotated_frame: np.ndarray, 
        violation_count: int, 
        all_detections: list[Detection], 
        db_collection: AsyncIOMotorCollection
    ):
    """Background task for saving violation record"""

    try:
        cloud_url = await upload_image_to_cloudinary(annotated_frame)
        print(f"Image uploaded to Cloudinary: {cloud_url}")

        timestamp_obj = datetime.now()

        violation_doc = ViolationRecord(
            timestamp=timestamp_obj,
            image_url=cloud_url,
            total_violations=violation_count,
            detections=[d for d in all_detections if d.class_id == 1]  # Chỉ lưu detections vi phạm (không đội mũ)
        )
        await db_collection.insert_one(violation_doc.model_dump())
        print(f"[Background] Saved violation to Cloud")
    except Exception as e:
        print(f"[Background] Failed to save history: {e}")